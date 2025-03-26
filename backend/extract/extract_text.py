import fitz  # PyMuPDF
import openpyxl
import pytesseract
from PIL import Image
from pdf2image import convert_from_path
from docx import Document
import os
import re
import tempfile

def extract_text_from_pdf(pdf_path):
    text = ""
    with fitz.open(pdf_path) as doc:
        for page in doc:
            text += page.get_text()
    if not text.strip():
        images = convert_from_path(pdf_path)
        for img in images:
            text += pytesseract.image_to_string(img)
    return text.strip()

def extract_text_from_docx(docx_path):
    doc = Document(docx_path)
    return "\n".join([para.text for para in doc.paragraphs if para.text.strip()])

def extract_text_from_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            text += " ".join(str(cell) for cell in row if cell) + "\n"
    return text.strip()

def clean_text(text):
    text = text.replace("\n", " ").strip()
    text = re.sub(r'\s+', ' ', text)
    return text

def extract_text_from_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.pdf':
        return clean_text(extract_text_from_pdf(file_path))
    elif ext in ['.doc', '.docx']:
        return clean_text(extract_text_from_docx(file_path))
    elif ext in ['.xls', '.xlsx']:
        return clean_text(extract_text_from_excel(file_path))
    else:
        raise ValueError("Unsupported file format")

def extract_text_from_fileobj(file_obj):
    ext = os.path.splitext(file_obj.name)[1].lower()
    file_obj.seek(0)  # Đảm bảo con trỏ về đầu file

    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as temp_file:
        temp_file.write(file_obj.read())
        temp_file_path = temp_file.name  # Lưu tên trước khi đóng

    # Sau khi đóng file tạm, tiếp tục extract
    text = extract_text_from_file(temp_file_path)

    # Xóa file tạm sau khi dùng xong
    os.remove(temp_file_path)

    return text

